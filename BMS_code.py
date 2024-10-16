import os
import pandas as pd
import datetime as dt
import openpyxl
from openpyxl import Workbook, load_workbook

book = load_workbook('BMS2.xlsx')
book = openpyxl.Workbook()
book = openpyxl.load_workbook("BMS2.xlsx")
sheet = book.active
sheet.title = "Bakery Management System"

# date/time module..
date = dt.datetime.now()
m = date.month
d = date.day
y = date.year

# logo..
bakery = "Bakery Management System"
bakery.center(45)
print(bakery)

# intro..
print("1. Add Order")
print("2. View Order")
print("3. Update Order")
print("4. Save to Excel")
print("5. Clear Data")
print("6. Exit")
ch1 = input("Enter Your Choice:-")

df = pd.read_excel("BMS2.xlsx")

if (ch1=="1"):
     sheet['A1'] = 'Customer Name'
     sheet['B1'] = 'Item'
     sheet['C1'] = 'Quantity'
     sheet['D1'] = 'Date'
# row counter..
     row = 5
     while True:
    # for user input..
       customer_name = input("Enter Customer Name (or type 'exit' to quit): ")
       if customer_name.lower() == 'exit':
         break
       item = input("Enter item Name: ")
       quantity = input("Enter quantity: ")
       date =  (f"{m}/{d}/{y}")

    # data to the Excel sheet
       sheet[f'A{row}'] = customer_name
       sheet[f'B{row}'] = item
       sheet[f'C{row}'] = quantity
       sheet[f'D{row}'] = date

       row += 1
     book.save('BMS2.xlsx')

elif (ch1 == "2"):
   book.save('BMS2.xlsx')
   df = pd.read_excel("BMS2.xlsx")
   print(df)

if (ch1=="3"):
     new_values = []
     for i in range(len(new_values)):
          sheet.cell(row=1, column=i+1).value = new_values[i]
     print("Update Completed...")

elif(ch1=="4"):
     book.save('BMS2.xlsx')
     print("Saved...")

elif (ch1=="5"):
    def clear_data():
        for row in sheet['A2:D' + str(sheet.max_row)]:
         for cell in row:
            cell.value = None
        print("All data cleared.")
    row = sheet.max_row + 1

    while True:
    # Ask for user input
       action = input("Enter 'add' to add entry, 'clear' to clear data, or 'exit' to quit: ")
    
       if action.lower() == 'exit':
         break
       elif action.lower() == 'clear':
          clear_data()
          row = 2  # Reset row counter after clearing data
          continue
       elif action.lower() != 'add':
         print("Invalid option. Please enter 'add', 'clear', or 'exit'.")
         continue
    book.save('BMS2.xlsx')
    


       
   






    
    
     









